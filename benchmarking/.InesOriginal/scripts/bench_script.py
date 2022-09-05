import numpy as np
import cmd 
import os 
import pandas as pd
import re 
import subprocess
import file_read_backwards
from file_read_backwards import FileReadBackwards
import time
from openpyxl import Workbook
import csv
import argparse
import datetime

def benchmark_single_result(file_path):
    '''
    function to convert proto buf binary to a readable dictionnary 
    message that would be stored later to the array
    '''
    with open(file_path, 'rb') as f:
        benchmark_result = test_log_pb2.BenchmarkEntries()    
        benchmark_result.ParseFromString(f.read())
    return benchmark_result       

def run_tf_cpu_gpu_benchmark(bench_folder, outcsv_cpugpu, accumulate):
    '''
    This function will call tensorflow benchmark for the different models we want to run on the cpu and gpu.
    This requires that tensorflow is built from source (tested and verified with tensorflow r1.13 and r1.14). 
    '''
    try:
        bench_folder = os.getcwd() +"/" + bench_folder 
        #bench_folder = os.getcwd() + "/" + input('path to folder to benchmark: ')
        test_parameters = {"list_of_threads":[1, 2], "list_of_runs":[1, 10], "list_of_batches":[1, 3, 112], "list_of_channels":[3],"list_of_shapes":["192, 192", "224, 224"]}
        columns = ["benchmark_name", "Hardware" , "num_threads" ,"input_type", "num_of_batches", "input_shape", "num_of_channels", "count", "first", "curr", "min", "max", "avg", "std"]
        
        if (accumulate == "new tests"): 
            df= pd.DataFrame(columns=columns)
        else:
            df = pd.read_csv(accumulate)
        
        #capture_regex is used as a pattern to search a string for that pattern. The concerned string is outputed by the benchmark. 
        capture_regex = ('Timings \(microseconds\): count=(.*) first=(.*) curr=(.*) min=(.*) max=(.*) avg=(.*) std=(.*)')
        
        for dirname in os.listdir(bench_folder):
            for  filename in os.listdir(bench_folder + "/" + dirname):
                if filename.endswith('pb'):
                    benchmark_name = filename[0:len(filename)-3]
                    graph_path = "../" + bench_folder + "/" + dirname + '/' + filename
                    graph_sum = summarize_graph(graph_path)
                    input_layer_string = graph_sum[0].split("=")[1]
                    input_layer_type_string = graph_sum[1].split("=")[1]
                    #NHWC ( number of batches , height , width and Channels )
                    input_layer_shape_string= graph_sum[2].split("=")[1]
                    
                    if not(":" in input_layer_shape_string) :
                        test_parameters["list_of_shapes"] = [input_layer_shape_string[input_layer_shape_string.find(",") + 1:input_layer_shape_string.rfind(",")]]
                        test_parameters["list_of_channels"] = [input_layer_shape_string[input_layer_shape_string.rfind(",") + 1:]]

                    output_layer_string = graph_sum[3].split("=")[1]
                    log_folder =  bench_folder + "/" + dirname + "/logfiles"
                    if not(os.path.exists(log_folder)):
                        subprocess.run("mkdir "+ "../" + bench_folder + "/" + dirname + "/logfiles", shell=True)
                    
                    for num_threads in test_parameters["list_of_threads"]:
                        for num_runs in test_parameters["list_of_runs"]:
                            for num_of_batches in test_parameters["list_of_batches"]:
                                for num_of_channels in test_parameters["list_of_channels"]:
                                    for shape in test_parameters["list_of_shapes"]:
                                        input_layer_shape_string = str(num_of_batches)+ "," + shape +"," + str(num_of_channels)
                                        parameters = str(num_threads)+ str(num_runs) + str(num_of_batches) + str(num_of_channels) + shape[:shape.find(",")]
                                        cmd_benchmark = " bazel-bin/tensorflow/tools/benchmark/benchmark_model  "+ " --graph=" + graph_path  + " --input_layer="+ input_layer_string+ " --input_layer_shape="+ input_layer_shape_string +  " --input_layer_type=" + input_layer_type_string+ " --output_layer=" + output_layer_string + " --num_threads="+ str(num_threads) + " --max_num_runs="+ str(num_runs) + " 2> " + "../" + log_folder + "/" + benchmark_name + "_" + str(parameters) + ".log"
                                        subprocess.run(cmd_benchmark, shell=True, cwd = "/tensorflow")
                                        log_file_path = log_folder +  "/" + benchmark_name + "_" + parameters + ".log"
                                        with FileReadBackwards(log_file_path, encoding="utf-8") as frb:
                                            for line in frb:
                                                match = re.search(capture_regex, line)
                                                if match!=None: 
                                                    df_r= pd.DataFrame({"benchmark_name":[benchmark_name], "Hardware": ["Desktop(CPU + GPU)"], "num_threads" : [num_threads], "input_type":[input_layer_type_string],
                                                    "num_of_batches": [num_of_batches], "input_shape": [shape], "num_of_channels":[num_of_channels],
                                                    "count": [match.group(1)], "first": [match.group(2)], "curr": [match.group(3)], "min": [match.group(4)],
                                                    "max": [match.group(5)], "avg":[match.group(6)] , "std":[match.group(7)]})
                                                    print(line)
                                                    df = df.append(df_r, ignore_index=True)
                                                    break                                                    
                continue
        df.to_csv(os.getcwd() + "/results/" + outcsv_cpugpu + ".csv") 
    finally:
        df.to_csv(os.getcwd() + "/results/" + outcsv_cpugpu + ".csv") 

    return os.getcwd() + "/results/" + outcsv_cpugpu + ".csv"

def run_tf_edgetpu_benchmark(bench_folder, outcsv_coral, accumulate):
    '''
    This function will call tensorflow benchmark for the different models we want to run on edge tpu.
    This requires that tensorflow is built from source (tested and verified with tensorflow r1.13 and r1.14). 
    '''
    try:
        bench_folder = os.getcwd() +"/" + bench_folder
        #bench_folder = os.getcwd() + "/" + input('path to folder to benchmark: ')
        test_parameters = {"list_of_threads":[1, 2, 3], "list_of_runs":[1, 10], "list_of_batches":[1], "list_of_channels":[3],"list_of_shapes":["224, 224"]}
        columns = ["benchmark_name", "Hardware" , "num_threads" ,"input_type", "num_of_batches", "input_shape", "num_of_channels", "count", "first", "curr", "min", "max", "avg", "std"]
        if (accumulate == "new tests"): 
            df= pd.DataFrame(columns=columns)
        else:
            df = pd.read_csv(os.getcwd() +"/" + accumulate)
        capture_regex = ('Timings \(microseconds\): count=(.*) first=(.*) curr=(.*) min=(.*) max=(.*) avg=(.*) std=(.*)')
        for dirname in os.listdir(bench_folder):
            for  filename in os.listdir(bench_folder + "/" + dirname):
                if filename.endswith('.pb'):
                    benchmark_name = filename[0:len(filename)-4]
                    graph_sum = summarize_graph(bench_folder + '/' + dirname + '/' + filename)
                    input_layer_string = graph_sum[0].split("=")[1]
                    input_layer_type_string = graph_sum[1].split("=")[1]
                    input_layer_shape_string= graph_sum[2].split("=")[1]
                    if "-1" in input_layer_shape_string:
                        input_layer_shape_string = input_layer_shape_string.replace("-1", "1")

                    if not(":" in input_layer_shape_string):
                        test_parameters["list_of_shapes"] = [input_layer_shape_string[input_layer_shape_string.find(",") + 1:input_layer_shape_string.rfind(",")]]
                        test_parameters["list_of_channels"] = [input_layer_shape_string[input_layer_shape_string.rfind(",") + 1:]]

                    output_layer_string = graph_sum[3].split("=")[1]

                    cmd_toco = "toco --graph_def_file="  + filename + " --output_file=" + benchmark_name[:-4] + ".tflite" + " --input_format=TENSORFLOW_GRAPHDEF --output_format=TFLITE --inference_type=QUANTIZED_UINT8 --input_shape=" +input_layer_shape_string + " --input_array=" + input_layer_string +" --output_array="+ output_layer_string +" --std_dev_values=127 --mean_values=127 --default_ranges_min=0 --default_ranges_max=100" 
                    start_time = time.time()
                    
                    subprocess.run(cmd_toco, shell=True, cwd="../" + bench_folder + "/" + dirname + "/")

                    cmd_edge = "~/../usr/bin/edgetpu_compiler -o "+ bench_folder + "/" + dirname + "  "+ bench_folder + "/" + dirname + "/" + benchmark_name[:-4] + ".tflite" 
                    
                    try:
                        os.system(cmd_edge)
                        extra_time = time.time() - start_time
                    except:
                        print('aborted , no edge tpu rep')
                    graph_path = bench_folder + "/" + dirname + '/' + benchmark_name[:-4] + "_edgetpu.tflite" 
                    log_folder =  bench_folder + "/" + dirname + "/logfiles"
                    if not(os.path.exists(log_folder)):
                        subprocess.run("mkdir "+ bench_folder + "/" + dirname + "/logfiles", shell=True)
                    
                    for num_threads in test_parameters["list_of_threads"]:
                        for num_runs in test_parameters["list_of_runs"]:
                            for num_of_batches in test_parameters["list_of_batches"]:
                                for num_of_channels in test_parameters["list_of_channels"]:
                                    for shape in test_parameters["list_of_shapes"]:
                                        input_layer_shape_string = str(num_of_batches)+ "," + shape +"," + str(num_of_channels)
                                        parameters = str(num_threads)+ str(num_runs) + str(num_of_batches) + str(num_of_channels) + shape[:shape.find(",")]
                                        cmd_benchmark = " bazel-bin/tensorflow/lite/tools/benchmark/benchmark_model  "+ " --graph=" + graph_path  + " --use_edgetpu=true" + " --num_threads="+ str(num_threads) + " --num_runs="+ str(num_runs) + " --enable_op_profiling=true " + " 2> " + log_folder + "/" + benchmark_name + "_" + str(parameters) + ".log"
                                        subprocess.run(cmd_benchmark, shell=True, cwd = "../tensorflow")
                                        log_file_path = log_folder +  "/" + benchmark_name + "_" + parameters + ".log"
                                        with FileReadBackwards(log_file_path, encoding="utf-8") as frb:
                                            for line in frb:
                                                match = re.search(capture_regex, line)
                                                print(line)
                                                if match!=None: 
                                                    df_r= pd.DataFrame({"benchmark_name":[benchmark_name], "Hardware": ["coral edgetpu"], "num_threads" : [num_threads], "input_type":["uint8"],
                                                    "num_of_batches": [num_of_batches], "input_shape": [shape], "num_of_channels":[num_of_channels],
                                                    "count": [match.group(1)], "first": [match.group(2)], "curr": [match.group(3)], "min": [match.group(4)],
                                                    "max": [match.group(5)], "avg":[match.group(6)] , "std":[match.group(7)]})
                                                    print(line)
                                                    df = df.append(df_r, ignore_index=True)
                                                    break                                                    
                continue
        df.to_csv( os.getcwd() + "/results/" + outcsv_coral + ".csv")  

    finally:
        df.to_csv(os.getcwd() + "/results/" + outcsv_coral + ".csv")
    
    return os.getcwd() + "/results/" + outcsv_coral + ".csv"
def run_tf_ncs_benchmark(bench_folder, outcsv_ncs, accumulate):
    '''
    This function calls ncs benchmark for different models on the NCS.
    This requires that tensorflow is built from source (tested and verified with tensorflow r1.13 and r1.14). 
    '''
    try:
        bench_folder = os.getcwd() +"/" + bench_folder
        bench_folder = os.getcwd() + bench_folder
        max_num_of_shaves = 12 
        test_parameters = {"list_of_threads":[1], "list_of_runs":[1, 2], "list_of_batches":[1, 3, 112], "list_of_channels":[3],"list_of_shapes":["192 192", "224 224"], "list_of_shaves":[6, 12]}
        columns = ["benchmark_name", "Hardware" , "num_of_shaves", "num_threads" ,"input_type", "input_shape", "count", "inference time (ms)"]
        if (accumulate == "new tests"): 
            df= pd.DataFrame(columns=columns)
        else:
            df = pd.read_csv(os.getcwd() +"/" + accumulate)
        capture_regex = ('Total inference time\\s+(.*)')
        for dirname in os.listdir(bench_folder):
            for  filename in os.listdir(bench_folder + "/" + dirname):
                if filename.endswith('pb') and filename.find("quant")== -1:
                    benchmark_name = filename[0:len(filename)-3]
                    graph_path = bench_folder + "/" + dirname + '/' + filename
                    graph_sum = summarize_graph(graph_path)
                    input_layer_string = graph_sum[0].split("=")[1]
                    input_layer_type_string = graph_sum[1].split("=")[1]
                    input_layer_shape_string= graph_sum[2].split("=")[1]
                    if not(":" in input_layer_shape_string) :
                        input_layer_shape_string = input_layer_shape_string[input_layer_shape_string.find(",") + 1:input_layer_shape_string.rfind(",")]
                        test_parameters["list_of_shapes"] = [input_layer_shape_string.replace(",", " ")]                
                    output_layer_string = graph_sum[3].split("=")[1]
                    log_folder =  bench_folder + "/" + dirname + "/logfiles"
                    if not(os.path.exists(log_folder)):
                        subprocess.run("mkdir "+ "../" + bench_folder + "/" + dirname + "/logfiles", shell=True)
                    for num_shaves in test_parameters["list_of_shaves"]:
                        for num_runs in test_parameters["list_of_runs"]:
                            for shape in test_parameters["list_of_shapes"]:
                                parameters = str(num_shaves) + str(num_runs) + str(shape[:shape.find(" ")]) 
                                log_file_path = log_folder +  "/" + benchmark_name + "_" + parameters + "ncs"+ ".log"
                                cmd_benchmark_ncs = "( mvNCProfile " + graph_path + " -s " + str(num_shaves) + " -in " + input_layer_string + " -on " + output_layer_string + " --numofruns " + str(num_runs) + ") > " + log_file_path
                                subprocess.run(cmd_benchmark_ncs, shell=True)
                                with FileReadBackwards(log_file_path, encoding="utf-8") as frb:
                                    for line in frb:
                                        line_withoutspaces = line.strip(" ")
                                        match = re.search(capture_regex, line_withoutspaces)
                                        if match!=None: 
                                            df2 = pd.DataFrame({"benchmark_name": [benchmark_name], "Hardware": ["NCS2"], "input_type": [input_layer_type_string],
                                            "input_shape": [shape], "count": [num_runs], "number of shaves": [num_shaves], "inference time (ms)": [match.group(1)]})
                                            print(line)
                                            df = df.append(df2, ignore_index=True)
                                            break 


                continue
        df.to_csv(os.getcwd() + "/results/" + outcsv_ncs + ".csv")  
    finally:
        df.to_csv(os.getcwd() + "/results/" + outcsv_ncs + ".csv")
    return os.getcwd() + "/results/" + outcsv_ncs + ".csv"

def summarize_graph(graph_path):
    """
    This function returns command line string summarizing the model to input to the benchmark command line. 
    """
    cmd_sum = "bazel-bin/tensorflow/tools/graph_transforms/summarize_graph --in_graph=" + graph_path
    output = subprocess.run(cmd_sum, shell=True, cwd="../tensorflow", stdout=subprocess.PIPE)
    string_init = str(output)
    string_int = string_init.split("bazel run tensorflow/tools/benchmark:benchmark_model")
    print(string_init)
    sum_regex = "show_flops (.*)"
    match = re.search(sum_regex, string_int[1][0:len(string_int[1])-4])
    return match.group(1).split(" ")

def run_benchmark(bench_folder, outcsv_all, accumulate):
    '''
    This function will execute tf, edge and ncs benchmark combined. 
    '''
    try:
        bench_folder = os.getcwd() +"/" + bench_folder
        test_parameters = {"list_of_threads":[1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12], "list_of_runs":[50], "list_of_batches":[1], "list_of_channels":[3],"list_of_shapes":["224, 224"]}
        columns = ["benchmark_name", "Hardware" , "num_threads" ,"input_type", "num_of_batches", "input_shape", "num_of_channels", "count", "inference_time(us)", "extra_time_to_execute(us)"]
        if (accumulate == "new tests"): 
            df= pd.DataFrame(columns=columns)
        else:
            df = pd.read_csv(os.getcwd() +"/" + accumulate)
            print(df)
        
        #Patterns for regular experessions:
        capture_regex_tf = ('Timings \(microseconds\): count=(.*) first=(.*) curr=(.*) min=(.*) max=(.*) avg=(.*) std=(.*)')
        capture_regex_ncs_1 = ('Total inference time\\s+(.*)')
        capture_regex_ncs_2 = ('Time to Execute\\s+(.*) ')
        
        for dirname in os.listdir(bench_folder):
            for  filename in os.listdir(bench_folder + "/" + dirname):
                if filename.endswith('pb'):
                    benchmark_name = filename[0:len(filename)-3]
                    graph_path = "../" + bench_folder + "/" + dirname + '/' + filename
                    graph_sum = summarize_graph(graph_path)
                    input_layer_string = graph_sum[0].split("=")[1]
                    input_layer_type_string = graph_sum[1].split("=")[1]
                    #NHWC ( number of batches , height , width and Channels )
                    input_layer_shape_string= graph_sum[2].split("=")[1]
                    if not(":" in input_layer_shape_string) :
                        test_parameters["list_of_shapes"] = [input_layer_shape_string[input_layer_shape_string.find(",") + 1:input_layer_shape_string.rfind(",")]]
                        test_parameters["list_of_channels"] = [input_layer_shape_string[input_layer_shape_string.rfind(",") + 1:]]

                    output_layer_string = graph_sum[3].split("=")[1]
                    log_folder =  bench_folder + "/" + dirname + "/logfiles"
                    if not(os.path.exists(log_folder)):
                        subprocess.run("mkdir "+ "../" + bench_folder + "/" + dirname + "/logfiles", shell=True)

                    cmd_toco = "toco --graph_def_file="  + filename + " --output_file=" + benchmark_name[:-4] + ".tflite" + " --input_format=TENSORFLOW_GRAPHDEF --output_format=TFLITE --inference_type=QUANTIZED_UINT8 --input_shape=" +input_layer_shape_string + " --input_array=" + input_layer_string +" --output_array="+ output_layer_string +" --std_dev_values=127 --mean_values=127 --default_ranges_min=0 --default_ranges_max=100" 
                    #start_time = time.time()
                    
                    subprocess.run(cmd_toco, shell=True, cwd="../" + bench_folder + "/" + dirname + "/")

                    cmd_edge = "~/../usr/bin/edgetpu_compiler -o "+ bench_folder + "/" + dirname + "  "+ bench_folder + "/" + dirname + "/" + benchmark_name[:-4] + ".tflite" 
                    
                    try:
                        os.system(cmd_edge)
                        #extra_time = (time.time() - start_time) *1e6
                    except:
                        print('aborted , no edge tpu rep')
                    graph_path_edge = bench_folder + "/" + dirname + '/' + benchmark_name[:-4] + "_edgetpu.tflite" 
                    log_folder =  bench_folder + "/" + dirname + "/logfiles"
                    if not(os.path.exists(log_folder)):
                        subprocess.run("mkdir "+ bench_folder + "/" + dirname + "/logfiles", shell=True)
                    
                    for num_threads in test_parameters["list_of_threads"]:
                        for num_runs in test_parameters["list_of_runs"]:
                            for num_of_batches in test_parameters["list_of_batches"]:
                                for num_of_channels in test_parameters["list_of_channels"]:
                                    for shape in test_parameters["list_of_shapes"]:
                                        input_layer_shape_string = str(num_of_batches)+ "," + shape +"," + str(num_of_channels)
                                        parameters = str(num_threads)+ str(num_runs) + str(num_of_batches) + str(num_of_channels) + shape[:shape.find(",")]
                                        cmd_benchmark_tf = " bazel-bin/tensorflow/tools/benchmark/benchmark_model  "+ " --graph=" + graph_path  + " --input_layer="+ input_layer_string+ " --input_layer_shape="+ input_layer_shape_string +  " --input_layer_type=" + input_layer_type_string+ " --output_layer=" + output_layer_string + " --num_threads="+ str(num_threads) + " --max_num_runs="+ str(num_runs) + " 2> " + "../" + log_folder + "/" + benchmark_name + "_" + str(parameters) + ".log"
                                        subprocess.run(cmd_benchmark_tf, shell=True, cwd = "/tensorflow")
                                        log_file_path = log_folder +  "/" + benchmark_name + "_" + parameters + ".log"
                                        with FileReadBackwards(log_file_path, encoding="utf-8") as frb:
                                            for line in frb:
                                                match = re.search(capture_regex_tf, line)
                                                if match!=None: 
                                                    df_r= pd.DataFrame({"benchmark_name":[benchmark_name], "Hardware": ["Desktop(CPU + GPU)"], "num_threads" : [num_threads], "input_type":[input_layer_type_string],
                                                    "num_of_batches": [num_of_batches], "input_shape": [shape], "num_of_channels":[num_of_channels],
                                                    "count": [match.group(1)], "inference_time(us)":[match.group(6)] , "extra_time_to_execute(us)":[0]})
                                                    print(line)
                                                    df = df.append(df_r, ignore_index=True)
                                                    break  
                                        os.remove(log_file_path)
                                        parameters = str(num_threads) + str(num_runs) + str(shape[:shape.find(" ")]) 
                                        log_file_path = log_folder +  "/" + benchmark_name + "_" + parameters + "ncs"+ ".log"
                                        cmd_benchmark_ncs = "( mvNCProfile " + graph_path + " -s " + str(num_threads) + " -in " + input_layer_string + " -on " + output_layer_string + " --numofruns " + str(num_runs) + ") > " + log_file_path
                                        subprocess.run(cmd_benchmark_ncs, shell=True)
                                        with open(log_file_path, 'r') as f:
                                            for line in f:
                                                line_withoutspaces = line.strip(" ")
                                                match_2 = re.search(capture_regex_ncs_2, line_withoutspaces)
                                                if match_2!= None:
                                                    time_to_execute = match_2.group(1).replace(":", "")
                                                    time_to_execute = float(time_to_execute.replace("ms", ""))*(1e3)
                                                    print(time_to_execute)
                                                    break
                                        
                                        with FileReadBackwards(log_file_path, encoding="utf-8") as frb:
                                            for line in frb:
                                                line_withoutspaces = line.strip(" ")
                                                match = re.search(capture_regex_ncs_1, line_withoutspaces)
                                                if match!=None: 
                                                    df2 = pd.DataFrame({"benchmark_name": [benchmark_name], "Hardware": ["NCS2"],"num_threads" : [num_threads], "input_type": [input_layer_type_string],
                                                    "num_of_batches": [num_of_batches], "input_shape": [shape], "num_of_channels":[num_of_channels], "count": [num_runs], "inference_time(us)": [float(match.group(1))*1e3], "extra_time_to_execute(us)":[float(time_to_execute)]})
                                                    print(line)
                                                    df = df.append(df2, ignore_index=True)
                                                    break 
                                        os.remove(log_file_path)
                                        input_layer_shape_string = str(num_of_batches)+ "," + shape +"," + str(num_of_channels)
                                        parameters = str(num_threads)+ str(num_runs) + str(num_of_batches) + str(num_of_channels) + shape[:shape.find(",")]
                                        cmd_benchmark_edge = " bazel-bin/tensorflow/lite/tools/benchmark/benchmark_model  "+ " --graph=" + graph_path_edge  + " --use_edgetpu=true" + " --num_threads="+ str(num_threads) + " --num_runs="+ str(num_runs) + " --enable_op_profiling=true " + " 2> " + log_folder + "/" + benchmark_name + "_" + str(parameters) + ".log"
                                        subprocess.run(cmd_benchmark_edge, shell=True, cwd = "../tensorflow")
                                        log_file_path = log_folder +  "/" + benchmark_name + "_" + parameters + ".log"
                                        with FileReadBackwards(log_file_path, encoding="utf-8") as frb:
                                            for line in frb:
                                                match = re.search(capture_regex_tf, line)
                                                if match!=None: 
                                                    df_r= pd.DataFrame({"benchmark_name":[benchmark_name], "Hardware": ["coral edgetpu"], "num_threads" : [num_threads], "input_type":["uint8"],
                                                    "num_of_batches": [num_of_batches], "input_shape": [shape], "num_of_channels":[num_of_channels],
                                                    "count": [match.group(1)], "inference_time(us)":[match.group(6)] ,  "extra_time_to_execute(us)":[0]})
                                                    print(line)
                                                    df = df.append(df_r, ignore_index=True)
                                                    break    
                                        os.remove(log_file_path)                                                          
                
                continue
        
        df.to_csv(os.getcwd() + "/results/" + outcsv_all + ".csv") 
    finally :
        df.to_csv(os.getcwd() + "/results/" + outcsv_all + ".csv")  
    return os.getcwd() + "/results/" + outcsv_all + ".csv"

def run_main():
    '''
    main 
    '''
    parser = argparse.ArgumentParser(description="Processing the benchmark of tensorflow models")
  
    parser.add_argument("--bench_folder",type=str,default="",help="benchmark folder that is needed to be precised")
    parser.add_argument("--cpugpu_benchmark", type=bool, default=False, help="run benchmark for cpu and gpu " )
    parser.add_argument("--ncs_benchmark", type=bool, default=False, help="run benchmark for ncs")
    parser.add_argument("--coral_benchmark", type=bool, default=False, help="run benchmark for coral")
    parser.add_argument("--all_benchmark", type=bool, default=False, help="run benchmark for all the hw")
    parser.add_argument("--accumulate", type=str, default="new tests", help="new tests or accumulate new test results with old results ")
    flags, unparsed = parser.parse_known_args()
    
    wb = Workbook()

    if flags.cpugpu_benchmark == True :
        outcsv_cpugpu = "outcsv_cpugpu_" + datetime.datetime.now().strftime("%H%M%S")
        filepath_1 = run_tf_cpu_gpu_benchmark(flags.bench_folder, outcsv_cpugpu, flags.accumulate)
        ws1 = wb.create_sheet("cpu_gpu")
        with open(filepath_1, 'r') as f:
            for row in csv.reader(f):
                ws.append(row)
    if flags.coral_benchmark == True:
        outcsv_coral = "outcsv_coral_" + datetime.datetime.now().strftime("%H%M%S")
        filepath_2 = run_tf_edgetpu_benchmark(flags.bench_folder, outcsv_coral, flags.accumulate)
        ws2 = wb.create_sheet("coral_edge_tpu")
        with open(filepath, 'r') as f:
            for row in csv.reader(f):
                ws.append(row)
    if flags.ncs_benchmark == True:
        outcsv_ncs = "outcsv_ncs_" + datetime.datetime.now().strftime("%d_%h_%m")
        filepath_3 = run_tf_ncs_benchmark(flags.bench_folder, outcsv_ncs, flags.accumulate)
        ws3 = wb.create_sheet("ncs")
        with open(filepath, 'r') as f:
            for row in csv.reader(f):
                ws.append(row)
    if flags.all_benchmark == True:
        outcsv_all = "outcsv_all_" + datetime.datetime.now().strftime("%d_%h_%m") 
        filepath_all = run_benchmark(flags.bench_folder, outcsv_all, flags.accumulate)
        ws3 = wb.create_sheet("all_benchmark")
        with open(filepath, 'r') as f:
            for row in csv.reader(f):
                ws.append(row)
    #The following line additionnaly saves the output of the benchmark to an excel book. 
    wb.save("benchmark" + datetime.datetime.now().strftime("%d_%h_%m")  + ".xlsx")



if __name__ == "__main__":
    run_main()